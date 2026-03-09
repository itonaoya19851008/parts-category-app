import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
import sqlite3
import json
import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional

# ページ設定
st.set_page_config(
    page_title="部品カテゴリ分類アプリ",
    page_icon="📦",
    layout="wide"
)

# ── DB設定 ────────────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history.db")

def init_db():
    """DBとテーブルを初期化（初回起動時のみ作成）"""
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS aggregation_history (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            saved_at    TEXT    NOT NULL,
            filename    TEXT    NOT NULL,
            order_no    TEXT,
            item_no     TEXT,
            item_name   TEXT,
            summary_json TEXT   NOT NULL
        )
    """)
    con.commit()
    con.close()

def save_to_history(summary_df: pd.DataFrame, filename: str):
    """集計結果をDBに保存する（1実行 = 1レコード）"""
    con = sqlite3.connect(DB_PATH)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 代表情報（先頭行から取得）
    order_no  = str(summary_df['Order＃'].iloc[0])  if 'Order＃'  in summary_df.columns and len(summary_df) > 0 else ''
    item_no   = str(summary_df['Item＃'].iloc[0])   if 'Item＃'   in summary_df.columns and len(summary_df) > 0 else ''
    item_name = str(summary_df['Item名称'].iloc[0]) if 'Item名称' in summary_df.columns and len(summary_df) > 0 else ''
    # DataFrameをJSON文字列に変換
    summary_json = summary_df.to_json(force_ascii=False, orient='records')
    con.execute(
        "INSERT INTO aggregation_history (saved_at, filename, order_no, item_no, item_name, summary_json) VALUES (?,?,?,?,?,?)",
        (now, filename, order_no, item_no, item_name, summary_json)
    )
    con.commit()
    con.close()

def load_all_runs() -> list:
    """全履歴の一覧を返す（id, saved_at, filename, order_no, item_no, item_name）"""
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT id, saved_at, filename, order_no, item_no, item_name FROM aggregation_history ORDER BY id DESC"
    ).fetchall()
    con.close()
    return rows

def load_run_summary(run_id: int) -> pd.DataFrame:
    """指定IDの集計表をDataFrameで返す"""
    con = sqlite3.connect(DB_PATH)
    row = con.execute("SELECT summary_json FROM aggregation_history WHERE id=?", (run_id,)).fetchone()
    con.close()
    if row is None:
        return pd.DataFrame()
    return pd.read_json(row[0], orient='records')

def delete_run(run_id: int):
    """指定IDの履歴を削除する"""
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM aggregation_history WHERE id=?", (run_id,))
    con.commit()
    con.close()

# 起動時にDB初期化
init_db()

# ── タイトル ──────────────────────────────────────────────────────────────
st.title("📦 部品カテゴリ分類・集計アプリ")
st.markdown("---")

def parse_excel_structure(file_content: bytes, filename: str) -> pd.DataFrame:
    """
    Excelファイルの特殊な構造を解析してDataFrameに変換
    
    データ構造:
      行3: ORDER情報 (col2=Order#, col4=Order名)
      行4: ITEM情報  (col2=Item#,  col4=Item名称)
      行7: メインヘッダー (PARTS, PIECE, NAME OF PARTS, SIZE, Q'TY, MATERIAL, etc.)
      行8: サブヘッダー  (ITEM, PARENT ASSY NO, SERIAL, 使用, 手配)
      行9〜: データ行
    
    キー構造:
      SERIAL (行8・列3) = 図番 = 各部品/アセンブリの一意識別子
      PARTS  (行7・列4) = 部品番号 (アセンブリ内の通し番号)
      PIECE  (行7・列5) = 構成部材番号 (00=本体, 01,02...=構成部材)
    """
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            df_raw = pd.read_excel(file_content, header=None, engine='openpyxl')
        else:
            df_raw = pd.read_csv(file_content, header=None, encoding='utf-8-sig')
        
        data = df_raw.values.tolist()
        
        # ORDER情報の抽出（行3, インデックス2）
        order_name = None
        order_no = None
        if len(data) > 2:
            row2 = data[2]
            order_no   = str(row2[2]).strip() if len(row2) > 2 and pd.notna(row2[2]) else None
            order_name = str(row2[4]).strip() if len(row2) > 4 and pd.notna(row2[4]) else None
        
        # ITEM情報の抽出（行4, インデックス3）
        item_no = None
        item_name = None
        if len(data) > 3:
            row3 = data[3]
            item_no   = str(row3[2]).strip() if len(row3) > 2 and pd.notna(row3[2]) else None
            item_name = str(row3[4]).strip() if len(row3) > 4 and pd.notna(row3[4]) else None
        
        # ヘッダー行の取得（行7=インデックス6, 行8=インデックス7）
        header_row_idx = 6
        sub_header_row_idx = 7
        data_start_idx = 8  # 行9からデータ
        
        if len(data) <= data_start_idx:
            raise ValueError("ヘッダー行が見つかりません")
        
        main_headers = data[header_row_idx]    # 行7
        sub_headers  = data[sub_header_row_idx] # 行8
        
        # 列名を結合して生成
        # 行7が空で行8に値がある場合は行8の値を使用
        combined_headers = []
        for i, (mh, sh) in enumerate(zip(main_headers, sub_headers)):
            mh_str = str(mh).replace('\r','').replace('\n',' ').replace('\u3000',' ').strip() if pd.notna(mh) else ''
            sh_str = str(sh).replace('\r','').replace('\n',' ').replace('\u3000',' ').strip() if pd.notna(sh) else ''
            
            if mh_str and sh_str:
                # 例: "Q'TY" + "使用" → "Q'TY 使用"
                combined_headers.append(f"{mh_str} {sh_str}")
            elif mh_str:
                combined_headers.append(mh_str)
            elif sh_str:
                combined_headers.append(sh_str)
            else:
                combined_headers.append(f'Unnamed_{i}')
        
        # データ行を抽出（PARTS列=インデックス3に値がある行）
        data_rows = []
        for i in range(data_start_idx, len(data)):
            row = data[i]
            if not any(pd.notna(val) and str(val).strip() for val in row[:10]):
                continue
            if len(row) > 3 and pd.notna(row[3]) and str(row[3]).strip():
                data_rows.append(row)
        
        if not data_rows:
            raise ValueError("有効なデータ行が見つかりません")
        
        df = pd.DataFrame(data_rows, columns=combined_headers)
        
        # ORDER/ITEM情報を追加
        df['Order名'] = order_name
        df['Order＃'] = order_no
        df['Item＃'] = item_no
        df['Item名称'] = item_name
        
        # SERIAL列の確認（列3 = インデックス2 のデータ）
        # combined_headersでは 'SERIAL' のはず
        serial_col = None
        for col in df.columns:
            if 'SERIAL' in str(col).upper():
                serial_col = col
                break
        if serial_col is None:
            # フォールバック：インデックス2の列を使用
            if len(df.columns) > 2:
                df = df.rename(columns={df.columns[2]: 'SERIAL'})
                serial_col = 'SERIAL'
        if serial_col and serial_col != 'SERIAL':
            df = df.rename(columns={serial_col: 'SERIAL'})
        
        # PARENT ASSY NO列の確認
        parent_col = None
        for col in df.columns:
            if 'PARENT' in str(col).upper():
                parent_col = col
                break
        if parent_col and parent_col != 'PARENT ASSY NO':
            df = df.rename(columns={parent_col: 'PARENT ASSY NO'})
        
        # === MASS列を確実に数値へ変換 ===
        mass_col_name = None
        for col in df.columns:
            if 'MASS' in str(col).upper():
                mass_col_name = col
                break
        
        if mass_col_name:
            df[mass_col_name] = (
                df[mass_col_name]
                .astype(str)
                .str.strip()
                .str.replace(',', '', regex=False)
                .str.replace('kg', '', regex=False, case=False)
                .str.replace(' ', '', regex=False)
                .replace({'nan': None, '-': None, '': None, 'None': None})
            )
            df[mass_col_name] = pd.to_numeric(df[mass_col_name], errors='coerce').fillna(0.0)
            # 列名を統一
            if mass_col_name != 'MASS (Per_Piece)':
                df = df.rename(columns={mass_col_name: 'MASS (Per_Piece)'})
        
        # PIECE列を文字列2桁ゼロパディングに正規化（0→"00", 1→"01"）
        if 'PIECE' in df.columns:
            def normalize_piece(x):
                s = str(x).strip()
                if s in ('', 'nan', 'None'):
                    return ''
                try:
                    return str(int(float(s))).zfill(2)
                except:
                    return s
            df['PIECE'] = df['PIECE'].apply(normalize_piece)
        
        # SERIAL列を文字列に統一
        if 'SERIAL' in df.columns:
            df['SERIAL'] = df['SERIAL'].apply(
                lambda x: str(x).strip() if pd.notna(x) and str(x).strip() not in ('', 'nan', 'None') else ''
            )
        
        return df
    
    except Exception as e:
        st.error(f"ファイル読み込みエラー: {str(e)}")
        raise

def normalize_column_name(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    """
    列名のバリエーションに対応して列名を取得
    """
    for name in possible_names:
        for col in df.columns:
            if name.lower() in str(col).lower():
                return col
    return None

def get_column_value(row: pd.Series, df: pd.DataFrame, possible_names: List[str], default=None):
    """
    列名のバリエーションに対応して値を取得
    """
    col_name = normalize_column_name(df, possible_names)
    if col_name and col_name in row.index:
        val = row[col_name]
        if pd.notna(val):
            return str(val).strip()
    return default

def parse_weight(weight_val) -> float:
    """
    重量値を数値に変換（既にfloat/intの場合はそのまま返す）
    """
    if weight_val is None:
        return 0.0
    
    # 既に数値型の場合はそのまま返す（DataFrameでpd.to_numeric済みの場合）
    if isinstance(weight_val, (int, float)):
        if np.isnan(weight_val):
            return 0.0
        return float(weight_val)
    
    # 文字列の場合は変換
    weight_str = str(weight_val).strip()
    
    if not weight_str or weight_str in ('', '-', 'nan', 'None', 'NaN'):
        return 0.0
    
    try:
        return float(weight_str)
    except (ValueError, TypeError):
        pass
    
    try:
        # カンマ・単位除去して変換
        weight_str = weight_str.replace(',', '').replace('kg', '').replace('KG', '').replace(' ', '')
        match = re.search(r'(\d+\.?\d*)', weight_str)
        if match:
            return float(match.group(1))
    except (ValueError, TypeError):
        pass
    
    return 0.0

def parse_quantity(qty_str: Optional[str]) -> int:
    """
    数量文字列を整数に変換
    """
    if not qty_str or qty_str == '' or qty_str == '-':
        return 0
    try:
        # 数値以外の文字を除去
        qty_str = re.sub(r'[^\d]', '', str(qty_str))
        if qty_str:
            return int(float(qty_str))
    except:
        pass
    return 0

def check_material_pattern(material: str, patterns: List[str]) -> bool:
    """
    材質がパターンに一致するかチェック
    """
    if not material:
        return False
    material_upper = material.upper()
    for pattern in patterns:
        # ワイルドカード対応
        if '*' in pattern:
            pattern_re = pattern.replace('*', '.*')
            if re.search(pattern_re, material_upper):
                return True
        elif pattern in material_upper:
            return True
    return False

def check_size_thickness(size: Optional[str], min_thickness: float = 80.0) -> bool:
    """
    SIZE列に指定以上の厚み（t80以上など）が含まれるかチェック
    """
    if not size:
        return False
    size_str = str(size).upper()
    # t80, t100, 80t, 100t などのパターンを検索
    pattern = rf't\s*(\d+\.?\d*)'
    matches = re.findall(pattern, size_str)
    for match in matches:
        try:
            thickness = float(match)
            if thickness >= min_thickness:
                return True
        except:
            pass
    return False

def get_parts_single_weight(parts_data: pd.DataFrame, mass_col: str) -> float:
    """
    PARTS単品重量を取得する
    
    ロジック:
      PIECE=00のMASS > 0 → その値を使用（単品部品）
      PIECE=00のMASS = 0 → PIECE=01,02,...の合計を使用（溶接アセンブリ）
    
    背景:
      溶接構造物（MILL HOUSING等）はPIECE=00のMASS=0が入力されており、
      実際の重量はPIECE=01,02,...の各素材重量の合計である。
    """
    piece_00 = parts_data[parts_data['PIECE'] == '00']
    
    if not piece_00.empty:
        mass_00 = parse_weight(piece_00.iloc[0][mass_col])
        if mass_00 > 0:
            return mass_00  # 単品部品: PIECE=00の重量をそのまま使用
    
    # PIECE=00が0またはない → 構成部材の合計を使用
    non_zero_pieces = parts_data[parts_data['PIECE'] != '00']
    if not non_zero_pieces.empty:
        total = sum(parse_weight(row[mass_col]) for _, row in non_zero_pieces.iterrows())
        return total
    
    return 0.0


def classify_part(parts_data: pd.DataFrame) -> Tuple[Optional[str], str]:
    """
    1つの部品グループ（1つのSERIAL内のPARTS）をカテゴリに分類
    
    引数: parts_data = 1つのSERIAL×PARTSに属する全PIECE行
    戻り値: (カテゴリ, 判定理由)
    
    分類ロジック: 1〜12の順番で条件を評価
    """
    if parts_data.empty:
        return (None, 'データなし')
    
    # 列名を取得
    mass_col     = normalize_column_name(parts_data, ['MASS (Per_Piece)', 'MASS'])
    material_col = normalize_column_name(parts_data, ['MATERIAL'])
    name_col     = normalize_column_name(parts_data, ['NAME OF PARTS', 'NAME'])
    size_col     = normalize_column_name(parts_data, ['SIZE'])
    summary_col  = normalize_column_name(parts_data, ['SUMMARY'])
    supply_col   = normalize_column_name(parts_data, ['SUPPLY'])
    
    # PIECE=00行（部品ヘッダー）と構成部材行（PIECE≠00）を分離
    piece_00     = parts_data[parts_data['PIECE'] == '00']
    sub_pieces   = parts_data[parts_data['PIECE'] != '00']  # PIECE=01,02,...
    
    # 「PARTS単品重量」: PIECE=00のMASSが0の場合は構成部材合計で補完
    parts_weight = get_parts_single_weight(parts_data, mass_col) if mass_col else 0.0
    
    # PIECE=00の名称・SUMMARY・SUPPLYを取得（部品ヘッダーの情報）
    header_row = piece_00.iloc[0] if not piece_00.empty else parts_data.iloc[0]
    
    # 1. SUMMARY欄に「MS-」または「JIS」を含む ➔ 定義対象外
    if summary_col:
        summary = str(header_row.get(summary_col, '') or '').strip()
        if 'MS-' in summary.upper() or 'JIS' in summary.upper():
            return ('d', f'SUMMARYにMS-またはJISを含む（{summary}）（定義対象外）')
    
    # 2. SUPPLY欄に「ES, SU, ST, CS」を含むか？
    has_supply = False
    supply_val = str(header_row.get(supply_col, '') or '') if supply_col else ''
    if any(x in supply_val.upper() for x in ['ES', 'SU', 'ST', 'CS']):
        has_supply = True
    
    if has_supply:
        # 3. NAME OF PARTSに「Roll/Roller」を含み、かつ「BRG/Bearing」を含まないか？
        part_name = str(header_row.get(name_col, '') or '') if name_col else ''
        has_roll    = 'ROLL' in part_name.upper() or 'ROLLER' in part_name.upper()
        has_bearing = 'BRG'  in part_name.upper() or 'BEARING' in part_name.upper()
        
        if has_roll and not has_bearing:
            # Roll/Roller（BRGなし）→ ステップ5以降の重量・材質判定へそのまま流す
            pass
        else:
            # Roll以外、またはBRG/Bearingあり → 定義対象外
            return ('d', 'SUPPLY条件あり、Roll/Roller以外（またはBRG）のため対象外')
    
    # 3. 超大物: 総重量が3000kg以上か？
    if parts_weight >= 3000:
        return ('Ds', f'重量{parts_weight:.1f}kg（3000kg以上）')
    
    # 4. 材質分岐: グループ全体（全PIECE）にSPCC/SPHC/SS*が1つでも含まれるか？
    has_spcc_sphc_ss = False
    if material_col:
        for _, row in parts_data.iterrows():
            mat = str(row.get(material_col, '') or '')
            if check_material_pattern(mat, ['SPCC', 'SPHC', 'SS*']):
                has_spcc_sphc_ss = True
                break
    
    if has_spcc_sphc_ss:
        # ── YESルート ──
        # ③ 構成部材にSS系以外の材質があり、その構成単品の重量が300kg以上か？
        for _, row in sub_pieces.iterrows():
            mat = str(row.get(material_col, '') or '') if material_col else ''
            if mat and not check_material_pattern(mat, ['SPCC', 'SPHC', 'SS*']):
                pm = parse_weight(row[mass_col]) if mass_col else 0
                if pm >= 300:
                    return ('Ds', f'SS系だが異材質の単品300kg以上あり（{mat}）{pm:.1f}kg')
        # 該当なし → ⑦へ合流
    
    else:
        # ── NOルート ──
        # ⑤ 部品(PARTS)の総重量が400kg以上か？
        if parts_weight >= 400:
            return ('Ds', f'SS系を含まず総重量{parts_weight:.1f}kg≥400kg')
        
        # ⑥ 構成部材にSCM*/SF*を含み、かつ部品(PARTS)の総重量が300kg以上か？
        has_scm_sf = False
        for _, row in parts_data.iterrows():
            mat = str(row.get(material_col, '') or '') if material_col else ''
            if check_material_pattern(mat, ['SCM*', 'SF*']):
                has_scm_sf = True
                break
        if has_scm_sf and parts_weight >= 300:
            return ('Ds', f'SCM/SFを含み総重量{parts_weight:.1f}kg≥300kg')
        # 該当なし → ⑦へ合流
    
    # ── ⑦: YES/NOルート合流地点 ──
    # 構成部材にSS系以外の材質があり、その構成単品の重量が100kg以上か？
    for _, row in sub_pieces.iterrows():
        mat = str(row.get(material_col, '') or '') if material_col else ''
        if mat and not check_material_pattern(mat, ['SPCC', 'SPHC', 'SS*']):
            pm = parse_weight(row[mass_col]) if mass_col else 0
            if pm >= 100:
                return ('Dm', f'異材質の単品100kg以上あり（{mat}）{pm:.1f}kg')
    # PIECE=00単体部品（子部品なし）の場合も判定
    if sub_pieces.empty and material_col:
        mat = str(header_row.get(material_col, '') or '')
        if mat and not check_material_pattern(mat, ['SPCC', 'SPHC', 'SS*']):
            if parts_weight >= 100:
                return ('Dm', f'異材質（{mat}）単品{parts_weight:.1f}kg≥100kg')
    # ⑦も該当なし → 次のステップ（極厚板t80）へ
    
    # 5. 極厚板: SIZEにt80以上があり、かつ総重量300kg以上か？
    if size_col:
        for _, row in parts_data.iterrows():
            size = row.get(size_col, None)
            if check_size_thickness(size, 80.0):
                # 構成部材の重量に300kg以上のものがあるか？
                for _, row2 in parts_data.iterrows():
                    pm = parse_weight(row2[mass_col]) if mass_col else 0
                    if pm >= 300:
                        return ('Ds', f'構成部材にt80以上、重量{pm:.1f}kg≥300kg')
                break  # t80以上あるが300kg未満 → 9へ
    
    # 6. 大物: 総重量が500kg以上か？
    if parts_weight >= 500:
        return ('Dm', f'重量{parts_weight:.1f}kg（500kg以上）')
    
    # 7. 配管: 名称に「PIPE」または「PIPING」を含むか？
    if name_col:
        for _, row in parts_data.iterrows():
            nm = str(row.get(name_col, '') or '').upper()
            if 'PIPE' in nm or 'PIPING' in nm:
                return ('PD', f'名称にPIPE/PIPINGを含む（{row.get(name_col,"")}）')
    
    # 8. 軽量品: 総重量が50kg未満か？
    if parts_weight < 50:
        return ('PD', f'重量{parts_weight:.1f}kg（50kg未満）')
    
    # 9. 残部品 → 後で按分処理（前半50%=Dm、後半50%=De）
    return (None, f'該当条件なし（残部品）重量{parts_weight:.1f}kg')

def classify_all_parts(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    全行に対してカテゴリを付与し、明細リストと集計表を作成
    
    分類単位: SERIAL（図番）× PARTS番号 の組み合わせ
    ※ SERIALが同じPARTSグループ全体（PIECE=00〜N）を1部品として扱う
    
    戻り値: (明細リスト DataFrame, 集計表 DataFrame)
    """
    classification_results = []
    
    # SERIAL列の確認
    serial_col = 'SERIAL' if 'SERIAL' in df.columns else None
    if serial_col is None:
        st.error("SERIAL列が見つかりません。列名を確認してください。")
        st.write(list(df.columns))
        return None, None
    
    items = df['Item＃'].unique()
    
    for item_no in items:
        item_df = df[df['Item＃'] == item_no].copy()
        
        # SERIAL × PARTS の組み合わせをユニーク取得
        if serial_col in item_df.columns:
            groups = item_df.groupby([serial_col, 'PARTS'], sort=False)
        else:
            groups = item_df.groupby(['PARTS'], sort=False)
        
        for group_key, group_data in groups:
            if isinstance(group_key, tuple):
                serial_no, parts_no = str(group_key[0]), str(group_key[1])
            else:
                serial_no, parts_no = '', str(group_key)
            
            category, reason = classify_part(group_data)
            
            # このSERIAL×PARTSグループの全行に分類結果を付与
            for idx in group_data.index:
                classification_results.append({
                    '_idx': idx,
                    'Category': category,
                    '判定理由': reason
                })
    
    # 分類結果をDataFrameに変換してインデックスでマージ
    cls_df = pd.DataFrame(classification_results).set_index('_idx')
    df_classified = df.copy()
    df_classified['Category'] = cls_df['Category']
    df_classified['判定理由'] = cls_df['判定理由']
    
    # 残部品（Category=None）のみを按分処理（ステップ12に到達した部品）
    # ※ 'd'カテゴリ（定義対象外：MS-/JIS品、Roll/Rollerなし）は按分対象外
    remaining_mask = df_classified['Category'].isna()
    
    if remaining_mask.sum() > 0:
        # SERIAL×PARTSのユニークグループで按分
        remaining_serials = df_classified[remaining_mask][[serial_col, 'PARTS']].drop_duplicates()
        total_remaining = len(remaining_serials)
        half = total_remaining // 2
        
        dm_serials = set()
        de_serials = set()
        for i, (_, row) in enumerate(remaining_serials.iterrows()):
            key = (str(row[serial_col]), str(row['PARTS']))
            if i < half:
                dm_serials.add(key)
            else:
                de_serials.add(key)
        
        for idx, row in df_classified[remaining_mask].iterrows():
            key = (str(row[serial_col]), str(row['PARTS']))
            if key in dm_serials:
                df_classified.at[idx, 'Category'] = 'Dm'
                df_classified.at[idx, '判定理由'] = '残部品（按分：Dm）'
            elif key in de_serials:
                df_classified.at[idx, 'Category'] = 'De'
                df_classified.at[idx, '判定理由'] = '残部品（按分：De）'
    
    # カテゴリ順でソート（Ds→Dm→PD→De→d→定義対象外）
    category_order = {'Ds': 1, 'Dm': 2, 'PD': 3, 'De': 4, 'd': 5}
    df_classified['_cat_order'] = df_classified['Category'].map(category_order).fillna(99)
    df_classified = df_classified.sort_values(['_cat_order', serial_col, 'PARTS', 'PIECE'],
                                               ascending=True, na_position='last')
    df_classified = df_classified.drop('_cat_order', axis=1)
    
    # 明細リスト用の列を選択
    mass_col_name = 'MASS (Per_Piece)' if 'MASS (Per_Piece)' in df_classified.columns else \
                    normalize_column_name(df_classified, ['MASS'])
    
    detail_columns = ['Order名', 'Order＃', 'Item＃', 'Item名称', 'SERIAL', 'PARENT ASSY NO',
                      'PARTS', 'PIECE', 'NAME OF PARTS', 'SIZE', 'MATERIAL']
    if mass_col_name:
        detail_columns.append(mass_col_name)
    detail_columns.extend(['SUPPLY', 'SUMMARY', 'Category', '判定理由'])
    
    available_columns = [col for col in detail_columns if col in df_classified.columns]
    detail_df = df_classified[available_columns].copy()
    if mass_col_name and mass_col_name != 'MASS (Per_Piece)' and mass_col_name in detail_df.columns:
        detail_df = detail_df.rename(columns={mass_col_name: 'MASS (Per_Piece)'})
    
    # 集計表を作成
    summary_df = create_summary_table(df_classified, items)
    
    return detail_df, summary_df

def create_summary_table(df_classified: pd.DataFrame, items: List) -> pd.DataFrame:
    """
    Itemごとにカテゴリ別の部品数と重量を集計
    
    集計単位: SERIAL × PARTS（1つのSERIALグループ = 1部品として数える）
    重量:     PARTS単品重量（PIECE=00が0ならPIECE合計）× Q'TY（使用数量）
    """
    mass_col = 'MASS (Per_Piece)' if 'MASS (Per_Piece)' in df_classified.columns else \
               normalize_column_name(df_classified, ['MASS'])
    serial_col = 'SERIAL' if 'SERIAL' in df_classified.columns else None
    
    # Q'TY 使用列を取得（"Q'TY 使用" or "Q'TY"）
    qty_col = None
    for col in df_classified.columns:
        col_up = str(col).upper()
        if ("Q'TY" in col_up or "QTY" in col_up) and "使用" in col_up:
            qty_col = col
            break
    if not qty_col:
        # フォールバック: 最初のQ'TY列
        qty_col = normalize_column_name(df_classified, ["Q'TY", 'QTY', '使用'])
    
    if not mass_col:
        st.error("MASS列が見つかりません")
        return pd.DataFrame()
    
    aggregation_data = []
    categories = ['d', 'Ds', 'Dm', 'De', 'PD']
    
    for item_no in items:
        item_df = df_classified[df_classified['Item＃'] == item_no]
        if item_df.empty:
            continue
        item_info = item_df.iloc[0]
        
        category_stats = {cat: {'count': 0, 'weight': 0.0} for cat in categories}
        
        # SERIAL × PARTS のユニークグループで集計
        if serial_col and serial_col in item_df.columns:
            group_cols = [serial_col, 'PARTS']
        else:
            group_cols = ['PARTS']
        
        processed_groups = set()
        for _, row in item_df.iterrows():
            if serial_col:
                gkey = (str(row.get(serial_col, '')), str(row['PARTS']))
            else:
                gkey = (str(row['PARTS']),)
            
            if gkey in processed_groups:
                continue
            processed_groups.add(gkey)
            
            category = row['Category']
            if not category or category not in categories:
                continue
            
            # このグループの全PIECE行を取得
            if serial_col:
                grp = item_df[(item_df[serial_col].astype(str) == gkey[0]) &
                              (item_df['PARTS'].astype(str) == gkey[-1])]
            else:
                grp = item_df[item_df['PARTS'].astype(str) == gkey[-1]]
            
            # PARTS単品重量（PIECE=00が0ならPIECE合計で補完）
            parts_w = get_parts_single_weight(grp, mass_col)
            
            # Q'TY（使用数量）を取得 — PIECE=00行から
            qty = 1
            piece_00_rows = grp[grp['PIECE'] == '00']
            if not piece_00_rows.empty and qty_col and qty_col in piece_00_rows.columns:
                qty = parse_quantity(piece_00_rows.iloc[0][qty_col]) or 1
            
            category_stats[category]['count']  += 1
            category_stats[category]['weight'] += parts_w * qty
        
        row_data = {
            'Order名':  item_info['Order名'],
            'Order＃':  item_info['Order＃'],
            'Item＃':   item_no,
            'Item名称': item_info['Item名称']
        }
        for cat in categories:
            row_data[f'{cat}_部品数'] = category_stats[cat]['count']
            row_data[f'{cat}_重量']   = round(category_stats[cat]['weight'], 2)
        
        aggregation_data.append(row_data)
    
    return pd.DataFrame(aggregation_data)


def create_excel_output(detail_df: pd.DataFrame, summary_df: pd.DataFrame) -> BytesIO:
    """
    スクリーンショットの形式に合わせたExcel出力
    
    集計表レイアウト（隣接した2段ヘッダー）:
      行1: [Order名] [Order#] [Item#] [Item名称] [  d  ] [ Ds  ] [ Dm  ] [ De  ] [ PD  ]
      行2: (結合済み)                              [数][重] [数][重] [数][重] [数][重] [数][重]
      行3〜: データ
    
    列配置:
      A:Order名  B:Order#  C:Item#  D:Item名称
      E:d_部品数  F:d_重量
      G:Ds_部品数 H:Ds_重量
      I:Dm_部品数 J:Dm_重量
      K:De_部品数 L:De_重量
      M:PD_部品数 N:PD_重量
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        # ── シート1: 明細リスト ──────────────────────────────
        detail_df.to_excel(writer, sheet_name='明細リスト', index=False)
        detail_ws = writer.sheets['明細リスト']

        # ヘッダー行のスタイル
        header_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        for cell in detail_ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 列幅（列数分だけ適用）
        col_widths = [12, 10, 10, 30, 10, 8, 10, 8, 30, 20, 15, 12, 8, 12, 8, 50]
        for i, width in enumerate(col_widths):
            if i < len(detail_df.columns):
                detail_ws.column_dimensions[get_column_letter(i + 1)].width = width
        detail_ws.row_dimensions[1].height = 30

        # ── シート2: 集計表 ──────────────────────────────────
        summary_ws = writer.book.create_sheet('集計表')

        # カテゴリごとの背景色
        cat_colors = {
            'd':  'FFFF99',   # 黄
            'Ds': 'FFCCCC',   # 赤系
            'Dm': 'FFD9B3',   # オレンジ系
            'De': 'CCFFCC',   # 緑系
            'PD': 'CCE5FF',   # 青系
        }

        # 罫線スタイル
        thin  = Side(style='thin')
        thick = Side(style='medium')
        def thin_border():
            return Border(left=thin, right=thin, top=thin, bottom=thin)
        def thick_border():
            return Border(left=thick, right=thick, top=thick, bottom=thick)

        # ── 行1: 上段ヘッダー ──
        # A-D: 属性列（行1:行2 結合）
        attr_headers = ['Order名', 'Order＃', 'Item＃', 'Item名称']
        attr_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        for col_idx, label in enumerate(attr_headers, start=1):
            summary_ws.merge_cells(
                start_row=1, start_column=col_idx,
                end_row=2,   end_column=col_idx
            )
            cell = summary_ws.cell(row=1, column=col_idx)
            cell.value = label
            cell.font = Font(bold=True)
            cell.fill = attr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border()

        # E以降: カテゴリ列（2列ずつ）
        categories = ['d', 'Ds', 'Dm', 'De', 'PD']
        cat_start_col = 5  # E列

        for i, cat in enumerate(categories):
            c1 = cat_start_col + i * 2      # 部品数列
            c2 = cat_start_col + i * 2 + 1  # 重量列
            fill = PatternFill(fill_type='solid', fgColor=cat_colors[cat])

            # 行1: カテゴリ名（2列結合）
            summary_ws.merge_cells(
                start_row=1, start_column=c1,
                end_row=1,   end_column=c2
            )
            cell = summary_ws.cell(row=1, column=c1)
            cell.value = cat
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border()

            # 行2: 部品数 / 重量
            for col_idx, label in [(c1, '部品数'), (c2, '重量')]:
                cell = summary_ws.cell(row=2, column=col_idx)
                cell.value = label
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border()

        # 行1・行2の高さ
        summary_ws.row_dimensions[1].height = 22
        summary_ws.row_dimensions[2].height = 18

        # ── 行3〜: データ行 ──
        for row_idx, (_, data_row) in enumerate(summary_df.iterrows(), start=3):
            # 属性列（A-D）
            for col_idx, key in enumerate(['Order名', 'Order＃', 'Item＃', 'Item名称'], start=1):
                cell = summary_ws.cell(row=row_idx, column=col_idx)
                cell.value = data_row.get(key, '')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border()

            # カテゴリ列（E以降）
            for i, cat in enumerate(categories):
                c1 = cat_start_col + i * 2
                c2 = cat_start_col + i * 2 + 1
                fill = PatternFill(fill_type='solid', fgColor=cat_colors[cat] + '40')  # 薄め

                cell_count = summary_ws.cell(row=row_idx, column=c1)
                cell_count.value = data_row.get(f'{cat}_部品数', 0)
                cell_count.alignment = Alignment(horizontal='center', vertical='center')
                cell_count.border = thin_border()

                cell_weight = summary_ws.cell(row=row_idx, column=c2)
                cell_weight.value = data_row.get(f'{cat}_重量', 0.0)
                cell_weight.number_format = '#,##0.00'
                cell_weight.alignment = Alignment(horizontal='right', vertical='center')
                cell_weight.border = thin_border()

        # ── 列幅の設定 ──
        col_widths_summary = {
            1: 14,   # Order名
            2: 12,   # Order#
            3: 10,   # Item#
            4: 28,   # Item名称
        }
        # カテゴリ列
        for i in range(len(categories)):
            col_widths_summary[cat_start_col + i * 2]     = 8   # 部品数
            col_widths_summary[cat_start_col + i * 2 + 1] = 12  # 重量

        for col_idx, width in col_widths_summary.items():
            summary_ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ウィンドウ枠固定（ヘッダー2行 + 属性4列）
        summary_ws.freeze_panes = 'E3'

    output.seek(0)
    return output

# ── セッションステート初期化 ───────────────────────────────────────────────
for key, default in [('detail_df', None), ('summary_df', None),
                     ('excel_output', None), ('last_filename', '')]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── タブ構成 ───────────────────────────────────────────────────────────────
tab_main, tab_history, tab_flow = st.tabs(["📂 分類・集計", "🕒 過去の集計履歴", "📊 判定フローチャート"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: 分類・集計
# ══════════════════════════════════════════════════════════════════════════════
with tab_main:
    uploaded_file = st.file_uploader(
        "部品表ファイルをアップロードしてください",
        type=['xlsx', 'xls', 'csv'],
        help="Excel形式（.xlsx, .xls）またはCSV形式のファイルをアップロードできます"
    )

    if uploaded_file is not None:
        try:
            with st.spinner("ファイルを読み込んでいます..."):
                file_content = uploaded_file.read()
                df = parse_excel_structure(BytesIO(file_content), uploaded_file.name)

            st.success(f"✅ ファイル読み込み完了: {len(df)} 行のデータを読み込みました")

            # デバッグ情報
            with st.expander("🔍 デバッグ情報（列名・データ型確認）", expanded=False):
                st.write("**列名一覧:**")
                st.write(list(df.columns))
                st.write("**データ型 (dtypes):**")
                st.write(df.dtypes)
                mass_debug_col = next((c for c in df.columns if 'MASS' in str(c).upper()), None)
                if mass_debug_col:
                    st.write(f"**{mass_debug_col} のサンプル（先頭10件）:**")
                    st.write(df[mass_debug_col].head(10).tolist())
                    st.write(f"最大値: **{df[mass_debug_col].max():.2f} kg**　最小値: **{df[mass_debug_col].min():.2f} kg**")

            # データプレビュー（全行）
            with st.expander("📋 読み込んだデータのプレビュー（全行）", expanded=False):
                st.dataframe(df, use_container_width=True, height=500)

            # 分類・集計ボタン
            if st.button("🔍 分類・集計を実行", type="primary"):
                with st.spinner("部品を分類・集計しています..."):
                    detail_df, summary_df = classify_all_parts(df)
                    if detail_df is not None and summary_df is not None:
                        st.session_state.detail_df    = detail_df
                        st.session_state.summary_df   = summary_df
                        st.session_state.excel_output = create_excel_output(detail_df, summary_df)
                        st.session_state.last_filename = uploaded_file.name

                        # ── DBに履歴として保存 ──
                        try:
                            save_to_history(summary_df, uploaded_file.name)
                            st.success("✅ 分類・集計完了！（結果を履歴に保存しました）")
                        except Exception as db_err:
                            st.success("✅ 分類・集計完了！")
                            st.warning(f"⚠️ 履歴の保存に失敗しました: {db_err}")

        except Exception as e:
            st.error(f"❌ エラーが発生しました: {str(e)}")
            st.exception(e)

    else:
        st.info("👆 上記から部品表ファイルをアップロードしてください")

    # 分類結果の表示（セッションステートから取得）
    if st.session_state.detail_df is not None:
        st.markdown("---")
        st.success("✅ 分類・集計結果")

        st.subheader("📋 カテゴリ付き明細リスト")
        st.caption(f"全 {len(st.session_state.detail_df)} 行")
        st.dataframe(st.session_state.detail_df, use_container_width=True, height=600)

        st.subheader("📊 集計表（Itemごと）")
        st.dataframe(st.session_state.summary_df, use_container_width=True)

        st.subheader("📥 Excelファイルのダウンロード")
        st.download_button(
            label="📥 明細リスト・集計表をExcel形式でダウンロード",
            data=st.session_state.excel_output,
            file_name="部品カテゴリ分類結果.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: 過去の集計履歴
# ══════════════════════════════════════════════════════════════════════════════
with tab_history:
    st.subheader("🕒 過去の集計履歴")

    runs = load_all_runs()

    if not runs:
        st.info("まだ集計履歴がありません。「分類・集計」タブで実行すると自動的に保存されます。")
    else:
        # 履歴一覧テーブル
        runs_df = pd.DataFrame(runs, columns=['ID', '保存日時', 'ファイル名', 'Order#', 'Item#', 'Item名称'])
        st.dataframe(runs_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        col_sel, col_del = st.columns([3, 1])

        # 選択ボックス: "ID: 日時 | ファイル名" 形式
        run_options = {f"#{r[0]}  {r[1]}  |  {r[2]}  ({r[5]})": r[0] for r in runs}

        with col_sel:
            selected_label = st.selectbox("📌 表示する履歴を選択", options=list(run_options.keys()))

        selected_id = run_options[selected_label]

        # 削除ボタン
        with col_del:
            st.write("")  # 高さ調整
            st.write("")
            if st.button("🗑️ 選択した履歴を削除", type="secondary"):
                delete_run(selected_id)
                st.success(f"履歴 #{selected_id} を削除しました。")
                st.rerun()

        # 選択した履歴の集計表を表示
        hist_summary_df = load_run_summary(selected_id)

        if not hist_summary_df.empty:
            st.subheader(f"📊 集計表（履歴 #{selected_id}）")
            st.dataframe(hist_summary_df, use_container_width=True)

            # 選択した履歴のExcelをダウンロード
            # 明細は保存していないので集計表のみ単独シートで出力
            hist_excel = BytesIO()
            with pd.ExcelWriter(hist_excel, engine='openpyxl') as writer:
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter

                hist_ws = writer.book.create_sheet('集計表')
                # writer.bookのデフォルトシートを削除
                if 'Sheet' in writer.book.sheetnames:
                    del writer.book['Sheet']

                cat_colors = {'d':'FFFF99','Ds':'FFCCCC','Dm':'FFD9B3','De':'CCFFCC','PD':'CCE5FF'}
                thin = Side(style='thin')
                def tb():
                    return Border(left=thin,right=thin,top=thin,bottom=thin)

                attr_headers = ['Order名','Order＃','Item＃','Item名称']
                attr_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
                for ci, lbl in enumerate(attr_headers, 1):
                    hist_ws.merge_cells(start_row=1,start_column=ci,end_row=2,end_column=ci)
                    c = hist_ws.cell(row=1, column=ci)
                    c.value = lbl; c.font = Font(bold=True)
                    c.fill = attr_fill
                    c.alignment = Alignment(horizontal='center',vertical='center')
                    c.border = tb()

                categories = ['d','Ds','Dm','De','PD']
                cat_start = 5
                for i, cat in enumerate(categories):
                    c1 = cat_start + i*2; c2 = c1+1
                    fill = PatternFill(fill_type='solid', fgColor=cat_colors[cat])
                    hist_ws.merge_cells(start_row=1,start_column=c1,end_row=1,end_column=c2)
                    hc = hist_ws.cell(row=1,column=c1)
                    hc.value=cat; hc.font=Font(bold=True); hc.fill=fill
                    hc.alignment=Alignment(horizontal='center',vertical='center'); hc.border=tb()
                    for col_idx, lbl in [(c1,'部品数'),(c2,'重量')]:
                        sc = hist_ws.cell(row=2,column=col_idx)
                        sc.value=lbl; sc.font=Font(bold=True); sc.fill=fill
                        sc.alignment=Alignment(horizontal='center',vertical='center'); sc.border=tb()

                for ri, (_, dr) in enumerate(hist_summary_df.iterrows(), start=3):
                    for ci, key in enumerate(attr_headers,1):
                        c = hist_ws.cell(row=ri,column=ci)
                        c.value=dr.get(key,''); c.border=tb()
                    for i, cat in enumerate(categories):
                        c1=cat_start+i*2; c2=c1+1
                        cc=hist_ws.cell(row=ri,column=c1); cc.value=dr.get(f'{cat}_部品数',0); cc.border=tb()
                        cw=hist_ws.cell(row=ri,column=c2); cw.value=dr.get(f'{cat}_重量',0.0)
                        cw.number_format='#,##0.00'; cw.border=tb()

                # 列幅
                for ci, w in [(1,14),(2,12),(3,10),(4,28)]:
                    hist_ws.column_dimensions[get_column_letter(ci)].width = w
                for i in range(len(categories)):
                    hist_ws.column_dimensions[get_column_letter(cat_start+i*2)].width = 8
                    hist_ws.column_dimensions[get_column_letter(cat_start+i*2+1)].width = 12
                hist_ws.freeze_panes = 'E3'

            hist_excel.seek(0)

            # ファイル名を履歴から取得
            run_info = next((r for r in runs if r[0] == selected_id), None)
            base_name = run_info[1].replace(':', '-').replace(' ', '_') if run_info else str(selected_id)
            st.download_button(
                label=f"📥 履歴 #{selected_id} の集計表をExcelでダウンロード",
                data=hist_excel,
                file_name=f"集計履歴_{base_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: 判定フローチャート
# ══════════════════════════════════════════════════════════════════════════════
with tab_flow:
    st.subheader("📊 部品分類 判定フローチャート")
    st.caption("classify_part 関数の判定ロジック（ステップ 1〜9）")

    FLOWCHART_HTML = """
<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <script src="https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.min.js"></script>
  <style>
    body { font-family:'Segoe UI','Hiragino Sans',sans-serif; background:#fff; margin:0; padding:12px; }
    .legend { display:flex; flex-wrap:wrap; gap:8px; margin-bottom:14px; }
    .legend-item { display:flex; align-items:center; gap:5px; font-size:0.78rem; color:#334155; }
    .legend-box { width:18px; height:13px; border-radius:2px; border:1px solid rgba(0,0,0,.15); }
  </style>
</head>
<body>
<div class="legend">
  <div class="legend-item"><div class="legend-box" style="background:#fde68a"></div>d（定義対象外）</div>
  <div class="legend-item"><div class="legend-box" style="background:#fca5a5"></div>Ds（設計重量品）</div>
  <div class="legend-item"><div class="legend-box" style="background:#fdba74"></div>Dm（製作重量品）</div>
  <div class="legend-item"><div class="legend-box" style="background:#86efac"></div>De（残部品後半50%）</div>
  <div class="legend-item"><div class="legend-box" style="background:#93c5fd"></div>PD（配管・軽量）</div>
</div>
<div class="mermaid">
flowchart TD
    START([部品グループ SERIAL x PARTS]):::start

    S1{"1 SUMMARY に<br/>MS- または JIS を含む?"}
    R1([d: 規格品・小物]):::cat_d

    S2{"2 SUPPLY に<br/>ES / SU / ST / CS を含む?"}
    S2b{"NAME に ROLL/ROLLER あり<br/>かつ BRG/BEARING なし?"}
    R2([d: 支給品のため対象外]):::cat_d

    S3{"3 総重量 >= 3000 kg?"}
    R3([Ds: 超大物 3000kg以上]):::cat_ds

    S4{"4 グループ全体に<br/>SPCC / SPHC / SS系<br/>が1つでも含まれる?"}

    S4y{"4-YES 構成単品に<br/>SS系以外の材質があり<br/>単品重量 >= 300 kg?"}
    R4y([Ds: SS系+異材質単品300kg以上]):::cat_ds

    S4n5{"4-NO 部品総重量<br/>>= 400 kg?"}
    R4n5([Ds: SS系なし+総重量400kg以上]):::cat_ds
    S4n6{"4-NO SCM* / SF* があり<br/>かつ 総重量 >= 300 kg?"}
    R4n6([Ds: SCM/SF+総重量300kg以上]):::cat_ds

    S7{"7 合流: 構成単品に<br/>SS系以外の材質があり<br/>単品重量 >= 100 kg?"}
    R7([Dm: 異材質単品100kg以上]):::cat_dm

    S5{"5 SIZE に t80以上 があり<br/>かつ 総重量 >= 300 kg?"}
    R5([Ds: 極厚板+300kg以上]):::cat_ds

    S6{"6 総重量 >= 500 kg?"}
    R6([Dm: 大物 500kg以上]):::cat_dm

    S8{"7 NAME に<br/>PIPE / PIPING を含む?"}
    R8([PD: 配管部品]):::cat_pd

    S9{"8 総重量 < 50 kg?"}
    R9([PD: 軽量品]):::cat_pd

    S10([9 残部品: 前半50% Dm / 後半50% De]):::cat_de

    START --> S1
    S1 -- YES --> R1
    S1 -- NO --> S2
    S2 -- NO --> S3
    S2 -- YES --> S2b
    S2b -- "ROLL/ROLLER かつ BRGなし" --> S3
    S2b -- "それ以外" --> R2
    S3 -- YES --> R3
    S3 -- NO --> S4
    S4 -- YES --> S4y
    S4 -- NO --> S4n5
    S4y -- YES --> R4y
    S4y -- NO --> S7
    S4n5 -- YES --> R4n5
    S4n5 -- NO --> S4n6
    S4n6 -- YES --> R4n6
    S4n6 -- NO --> S7
    S7 -- YES --> R7
    S7 -- NO --> S5
    S5 -- YES --> R5
    S5 -- NO --> S6
    S6 -- YES --> R6
    S6 -- NO --> S8
    S8 -- YES --> R8
    S8 -- NO --> S9
    S9 -- YES --> R9
    S9 -- NO --> S10

    classDef start  fill:#1e293b,stroke:#1e293b,color:#fff
    classDef cat_d  fill:#fde68a,stroke:#d97706,color:#78350f,font-weight:bold
    classDef cat_ds fill:#fca5a5,stroke:#dc2626,color:#7f1d1d,font-weight:bold
    classDef cat_dm fill:#fdba74,stroke:#ea580c,color:#7c2d12,font-weight:bold
    classDef cat_de fill:#86efac,stroke:#16a34a,color:#14532d,font-weight:bold
    classDef cat_pd fill:#93c5fd,stroke:#2563eb,color:#1e3a8a,font-weight:bold
</div>
<script>
  mermaid.initialize({
    startOnLoad:true, theme:'base',
    themeVariables:{ fontSize:'13px' },
    flowchart:{ curve:'basis', nodeSpacing:36, rankSpacing:46, useMaxWidth:false }
  });
</script>
</body>
</html>
"""
    import streamlit.components.v1 as components
    components.html(FLOWCHART_HTML, height=1100, scrolling=True)

